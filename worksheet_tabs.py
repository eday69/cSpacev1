# Created by Eric Day
# Sept 12 2018

import json
# Show current directory
import os
cwd = os.getcwd()
print(cwd)

# File name to analize/import
excel_file='cSpace_Booking.xlsx'

# load module
import openpyxl
# Open file
wb = openpyxl.load_workbook(excel_file)

sheets=wb.sheetnames
num_sheets=len(sheets)
print(json.dumps (sheets))
print('We have '+str(num_sheets)+' worksheets tabs in the file')
for i in range(num_sheets):
  print('Tab # '+str(i+1)+' is '+sheets[i])

#Worksheets:
#    Rates
#    Facilities
#    Clients
#    Dates...open?

# Read Rates
# Read Facilities
# Read Clients
# Read the rest of sheets as dates  by date
#     separate by hour (range) x room (client)


#for row in ws.rows:
#    for cell in row:
#        print(cell.value)


#print(wb2.sheetnames)
