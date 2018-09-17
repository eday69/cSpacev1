
from flask import Flask
app = Flask(__name__)

import json

# File name to analize/import
excel_file='cSpace_Booking.xlsx'

def openfile(file, wb):
    # load module
    from openpyxl import load_workbook
    # open file
    workbook = load_workbook(excel_file)
    try:
        ws = workbook[wb]  # Do we have a 'Clients' worksheet
    except Exception as err: # No,
        print('An exception happened: ' + str(err))
        return 0
    return ws

def getBookingInfo(myfile):
    fullinfo={};
    # load module
    from openpyxl import load_workbook
    # open file
    workbook = load_workbook(excel_file)
    sheets=workbook.sheetnames
    num_sheets=len(sheets)
    bookings={}
    for i in range(num_sheets):
        if (sheets[i] not in ['Clients', 'Facilities', 'Rates']):
            month=sheets[i]
            ws = workbook[month]
            # print('Tab # '+str(i+1)+' is '+month)
            dataroom = {}
            for row in ws.iter_cols(min_col=3):
                room = row[0].value
                dataday = {}
                for cell in row[1:]:
                    dataday[cell.row-1] = cell.value;
                # print(dataday)
                dataroom[room] = dataday
            bookings[month]=dataroom
    fullinfo['bookings']=bookings;
    # print(bookings)
    rates = {}
    location = ""
    rate = {}
    ws = workbook['Rates']
    for row in ws.iter_rows(min_row=2):
        new_client=row[0].value.split(' ')
        location = row[0].value
        rate[row[1].value] = row[2].value
    rates[location] = rate;
    fullinfo['rates'] = rates;
    client = {}
    ws = workbook['Clients']
    for row in ws.iter_rows(min_row=2):
        new_client=row[0].value.split(' ')
        client[new_client[0]] = { new_client[0] : new_client[1], 'issues':row[5].value};
    fullinfo['clients'] = client;
    ws = workbook['Facilities']
    facility = {}
    for row in ws.iter_rows(min_row=2):
        facility[row[0].value] = {
            "Type" : row[1].value,
            "Location" : row[2].value,
            "Description" : row[3].value,
            "Issues" : row[4].value }
    fullinfo['facilities'] = facility;

    return json.dumps(fullinfo);
    # print(bookings.keys())



def getClientInfo(myfile):
    worksheet=openfile(myfile, 'Clients')
    clients = []
    if worksheet:
        import os
        mydir = os.getcwd()
    #    for row in worksheet.rows:
        for row in worksheet.iter_rows(min_row=2):
            new_client=row[0].value.replace(u'\xa0', u' ').split(' ')
            client = {}
            client["first_name"] = new_client[0]
            client["last_name"] = new_client[1]
            client["issues"] = ""
            issues=row[6].value
            if (row[6].value):
                client["issues"] = issues
            clients.append(client)
        return clients
#        print(json.dumps(clients, ensure_ascii=False, indent=4))

def getRateInfo(myfile):
    worksheet=openfile(myfile, 'Rates')
    rates = {}
    location = ""
    if worksheet:
        import os
        mydir = os.getcwd()
    #    for row in worksheet.rows:
        rate = {}
        for row in worksheet.iter_rows(min_row=2):
            new_client=row[0].value.replace(u'\xa0', u' ').split(' ')
            location = row[0].value
            rate[row[1].value] = row[2].value
        rates[location] = rate;
        return json.dumps(rates);
#        print(json.dumps(clients, ensure_ascii=False, indent=4))

def getFacilitiesInfo(myfile):
    worksheet=openfile(myfile, 'Facilities')
    facilities = []
    if worksheet:
        import os
        mydir = os.getcwd()
    #    for row in worksheet.rows:
        for row in worksheet.iter_rows(min_row=2):
            new_client=row[0].value.split(' ')
            facility = {}
            facility["Room Name"] = row[0].value
            facility["Type"] = row[1].value
            facility["Location"] = row[2].value
            facility["Description"] = row[3].value
            facility["Issues"] = row[4].value
            facilities.append(facility)
        return json.dumps(facilities);
#        print(json.dumps(clients, ensure_ascii=False, indent=4))

# getBookingInfo(excel_file)

def printObject(jobj):
    rtnStr="<table style='border:1px solid red'>"
    for v in jobj:
        rtnStr+="<tr>"
        rtnStr+="<td>"+v['first_name']+"</td>"
        rtnStr+="<td>"+v['last_name']+"</td>"
        rtnStr+="<td>"+v['issues']+"</td>"
        rtnStr+="</tr>"
    rtnStr+="</table>"
    return rtnStr

@app.route('/')
def root():
    return app.send_static_file('index.html')

@app.route('/images/cSpace.jpg')
def image1():
    return app.send_static_file('images/cSpace.jpg')

@app.route('/images/cSpace_logo.png')
def image_logo():
    return app.send_static_file('images/cSpace_logo.png')

@app.route('/cspacebookings.html')
def cspacebookings():
    print('Here')
    return app.send_static_file('cspacebookings.html')


@app.route("/Clients")
def clientInfo():
    # wf = openfile(excel_file)
    ci = getClientInfo(excel_file)
    return printObject(ci)
#    return json.dumps(ci, ensure_ascii=False, indent=4)

@app.route("/Rates")
def rateInfo():
    # wf = openfile(excel_file)
    ci = getRateInfo(excel_file)
    return ci

@app.route("/Facilities")
def facilityInfo():
    print('In wrong place')
    ci = getFacilitiesInfo(excel_file)
    return ci

@app.route("/bookinfo")
def bookingsInfo():
    print('In right place')
    ci = getBookingInfo(excel_file)
    return ci
