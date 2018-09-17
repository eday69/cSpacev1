# Created by Eric Day
# Sept 12 2018

import json

# File name to analize/import
excel_file='cSpace Booking.xlsx'

def openfile(file):
    # load module
    from openpyxl import load_workbook
    # open file
    workbook = load_workbook(excel_file)
    try:
        ws = workbook['Clients']  # Do we have a 'Clients' worksheet
    except Exception as err: # No,
        print('An exception happened: ' + str(err))
        return 0
    return ws


worksheet=openfile(excel_file)
if worksheet:
    import os
    mydir = os.getcwd()
    myfile = open(mydir + '//test.html', 'w')
    myfile.write('<h1>Hello Tables</h1>')
    myfile.write('<table style=\'border: 2px solid green\'>')
#    for row in worksheet.rows:
    clients = []
    for row in worksheet.iter_rows(min_row=2):
        new_client=row[0].value.replace(u'\xa0', u' ').split(' ')
        client = {}
        client["first_name"] = new_client[0]
        client["last_name"] = new_client[1]
        client["issues"] = ""
        issues=row[6].value
        if (row[6].value):
            client["issues"] = issues
        clients.append(client)
        myfile.write('<tr><td>'+new_client[0]+'</td><td>'+new_client[1]+'</td><td>'+str(issues)+'</td>')
        myfile.write('</tr>')

    myfile.write('</table>')
    myfile.close()
    print(json.dumps(clients, ensure_ascii=False, indent=4))







#Worksheets:
#    Rates
#    Facilities
#    Clients
#    Dates...open?

# Read Rates
# Read Facilities
# Read Clients
# Read the rest of sheets as dates  by date
#     separate by hour (range) x room (client)




#print(wb2.sheetnames)
